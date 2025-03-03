import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import aiohttp
import asyncio
import pandas as pd
from datetime import datetime
from requests.auth import HTTPBasicAuth
import logging
from aiohttp import BasicAuth, ClientSession, ClientTimeout
from tqdm.asyncio import tqdm
import os
import getpass
from typing import Any, Dict, Optional, List
import psutil
import sqlite3

# ------------------------------------------------------------------------------
# Logging Setup
# ------------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('app.log', mode='w')]
)

# ------------------------------------------------------------------------------
# Simple color-coded print helper
# ------------------------------------------------------------------------------
class color:
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

# ------------------------------------------------------------------------------
# Comparison logic for previous runs
# ------------------------------------------------------------------------------
def compare_with_previous_run(
    current_data: pd.DataFrame,
    previous_file: str
) -> pd.DataFrame:
    # If no previous CSV, just return current with no 'Change' flagged
    if not os.path.exists(previous_file):
        current_data['Change'] = ''
        return current_data

    # Load previous data
    previous_data = pd.read_csv(previous_file, dtype={'Код товара': str})

    # Ensure both columns are strings for a consistent merge
    current_data['Код товара'] = current_data['Код товара'].astype(str)
    previous_data['Код товара'] = previous_data['Код товара'].astype(str)

    # Start by assuming no changes
    current_data['Change'] = ''

    # Prepare new 'Change' column for the previous dataset as well
    previous_data['Change'] = ''

    # Mark new rows
    new_mask = ~current_data['Код товара'].isin(previous_data['Код товара'])
    current_data.loc[new_mask, 'Change'] = 'New'

    # Mark disappeared rows
    disappeared_mask = ~previous_data['Код товара'].isin(current_data['Код товара'])
    # We'll keep the disappeared rows separately and append them with 'Disappeared'
    disappeared_rows = previous_data.loc[disappeared_mask].copy()
    disappeared_rows['Change'] = 'Disappeared'

    # For changed stock, merge on 'Код товара'
    merged_data = pd.merge(
        current_data[['Код товара', 'Остаток']],
        previous_data[['Код товара', 'Остаток']],
        on='Код товара',
        how='inner',
        suffixes=('_current', '_previous')
    )
    changed_mask = merged_data['Остаток_current'] != merged_data['Остаток_previous']
    changed_codes = merged_data.loc[changed_mask, 'Код товара'].unique()

    # Mark changed stocks in current_data
    current_data.loc[current_data['Код товара'].isin(changed_codes), 'Change'] = 'Stock Changed'

    # Combine current data + disappeared rows
    combined_data = pd.concat([current_data, disappeared_rows], ignore_index=True)

    return combined_data

# ------------------------------------------------------------------------------
# Highlight changes in the Excel sheet
# ------------------------------------------------------------------------------
def highlight_changes(ws, combined_data: pd.DataFrame):
    """
    Applies color highlighting in the Excel worksheet for changes.
    """
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # new
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # disappeared
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # stock changed

    for index, row in combined_data.iterrows():
        change_type = row.get('Change', '')
        if change_type == 'New':
            for cell in ws[index + 2]:
                cell.fill = green_fill
        elif change_type == 'Disappeared':
            for cell in ws[index + 2]:
                cell.fill = red_fill
        elif change_type == 'Stock Changed':
            for cell in ws[index + 2]:
                cell.fill = yellow_fill

# ------------------------------------------------------------------------------
# Formatting of excel sheet
# ------------------------------------------------------------------------------
def add_dropdown_and_formatting(file_path: str, sheet_name: str = 'Sheet1', start_row: int = 2, end_row: int = 2000) -> None:
    """
    Opens or creates an Excel workbook, adds dropdown menus, applies conditional formatting,
    sets column widths and general formatting.
    """
    dropdown_columns = ['F', 'G']
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        ws.title = sheet_name

    # Add dropdown menu
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=True)
    ws.add_data_validation(dv)

    # Apply dropdown to specified range
    for col in dropdown_columns:
        for row in range(start_row, end_row + 1):
            cell = f'{col}{row}'
            dv.add(ws[cell])

    # Apply conditional formatting for "Да" (green)
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(
            f'{col}{start_row}:{col}{end_row}',
            CellIsRule(operator='equal', formula=['"Да"'], stopIfTrue=True, fill=green_fill)
        )

    # Apply conditional formatting for "Нет" (orange)
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(
            f'{col}{start_row}:{col}{end_row}',
            CellIsRule(operator='equal', formula=['"Нет"'], stopIfTrue=True, fill=orange_fill)
        )
    
    # Apply auto-filter
    ws.auto_filter.ref = f"A1:Z{end_row}"

    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    for col in ['A', 'E', 'F', 'G', 'K']:
        ws.column_dimensions[col].width = 15

    # Apply additional formatting: general cell style and header style.
    for row in ws.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    wb.save(file_path)
    print(color.CYAN + f"Файл отформатирован {file_path}" + color.END)

# ------------------------------------------------------------------------------
# Auth info (adjust these or prompt user)
# ------------------------------------------------------------------------------
USERNAME = input("Enter your email: ")
PASSWORD = getpass.getpass("Enter your password: ")
auth = BasicAuth(USERNAME, PASSWORD)

base_url = "https://api.moysklad.ru/api/remap/1.2/entity/assortment"

# ------------------------------------------------------------------------------
# Fetching logic
# ------------------------------------------------------------------------------
MAX_REQUESTS = 5        # Limit concurrent requests
PAGE_SIZE = 1000        # MoySklad max page size
INCLUDED_PRICE_TYPES = [
    "Цена розница",
    "Цена маркетплейс",
    "Цена мелкий опт",
    "Цена средний опт"
]

async def fetch(
    session: ClientSession,
    url: str,
    retries: int = 5
) -> Optional[Dict[str, Any]]:
    for attempt in range(retries):
        try:
            async with session.get(url, auth=auth) as response:
                response.raise_for_status()
                return await response.json()

        except aiohttp.ClientResponseError as e:
            # 429: Too Many Requests
            if e.status == 429:
                logging.warning(f"Attempt {attempt + 1} got 429. Retrying...")
                if attempt < retries - 1:
                    retry_after = int(e.headers.get('Retry-After', 1))
                    await asyncio.sleep(retry_after)
                else:
                    logging.error(f"All {retries} attempts failed for {url} (429).")
                    return None
            else:
                logging.error(f"Request failed ({e.status}) for {url}: {e}")
                return None

        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            logging.warning(f"Attempt {attempt + 1} for {url} failed: {e}")
            if attempt < retries - 1:
                await asyncio.sleep(1)  # slight delay
            else:
                logging.error(f"All {retries} attempts failed for {url}")
                return None
    return None

async def fetch_product_details(
    session: ClientSession,
    product: Dict[str, Any],
    base_product_paths: Dict[str, str]
) -> Optional[Dict[str, Any]]:

    path_name = product.get('pathName', "")

    # If it's a base product, store its path
    if product.get('variantsCount', -1) > 0:
        base_product_paths[product['id']] = path_name

    # If it's a variant, get the parent product's path
    if product.get('meta', {}).get('type') == 'variant':
        parent_id = product.get('product', {}).get('meta', {}).get('href', '').split('/')[-1]
        path_name = base_product_paths.get(parent_id, "")

    # Attempt to extract "Категория" from characteristics if present
    category_value = "base"
    characteristics: List[Dict[str, Any]] = product.get('characteristics', [])
    for char in characteristics:
        category_value = str(char.get('value', 'base'))
        break

    # Filter only salePrices we care about
    sale_prices = product.get('salePrices', [])
    prices = {
        p.get('priceType', {}).get('name', 'Unknown'): (p.get('value', 0) / 100)
        for p in sale_prices
        if p.get('priceType', {}).get('name') in INCLUDED_PRICE_TYPES
    }

    return {
        'name': product.get('name'),
        'code': product.get('code'),
        'path': path_name,
        'stock': product.get('stock', 0),
        'days': product.get('stockDays', 0),
        'category': category_value,
        'prices': prices,
        'id': product.get('id')
    }

async def fetch_all_products(session: ClientSession, base_url: str, limit: int = PAGE_SIZE):
    """
    Fetch all products using pagination with a global progress bar.
    """
    offset = 0
    all_items = []
    base_product_paths = {}

    with tqdm(desc="Fetching Products", unit="batch", leave=False) as pbar:
        while True:
            url = f"{base_url}?limit={limit}&offset={offset}"
            logging.info(f"Fetching page offset={offset} ...")

            data = await fetch(session, url)
            if not data:
                logging.warning("No data returned, stopping pagination.")
                break

            rows = data.get('rows', [])
            if not rows:
                logging.info("No more rows; pagination complete.")
                break

            all_items.extend(rows)
            offset += limit
            pbar.update(1)  # Update progress bar for each batch

    return all_items, base_product_paths


# ------------------------------------------------------------------------------
# Check if the Excel file is open and prompt the user to close it
# ------------------------------------------------------------------------------
def check_and_prompt_close_excel(filename: str):
    while True:
        file_closed = True
        for proc in psutil.process_iter(['pid', 'name']):
            if proc.info['name'] == 'EXCEL.EXE':
                for open_file in proc.open_files():
                    if open_file.path == os.path.abspath(filename):
                        file_closed = False
                        print(color.RED + f"The file {filename} is currently open. Please close it." + color.END)
                        input("Press Enter after closing the file...")
                        break
            if not file_closed:
                break
        if file_closed:
            break

# ------------------------------------------------------------------------------
# Create or update the database
# ------------------------------------------------------------------------------
def update_database(data: pd.DataFrame, db_path: str):
    conn = sqlite3.connect(db_path)
    data.to_sql('products', conn, if_exists='replace', index=False)
    conn.close()
    print(color.GREEN + f"Data saved into database {db_path}" + color.END)
    
def chunkify(lst, chunk_size):
    """Splits a list into chunks of a given size."""
    for i in range(0, len(lst), chunk_size):
        yield lst[i:i + chunk_size]

# ------------------------------------------------------------------------------
# Main
# ------------------------------------------------------------------------------
async def main():
    filename = "all_products.xlsx"
    db_path = "all_products.db"
    check_and_prompt_close_excel(filename)
    previous_csv = "last.csv"
    CHUNK_SIZE = 1000

    # You can increase total=120 or so for large sets, but keep an eye on memory
    timeout = ClientTimeout(total=120)

    # Limit concurrency
    semaphore = asyncio.Semaphore(MAX_REQUESTS)

    async with ClientSession(timeout=timeout) as session:
        with tqdm(total=5, desc="Overall Progress", unit="step") as global_pbar:
            # 1) Fetch all products
            logging.info("Fetching list of all products...")
            products, base_product_paths = await fetch_all_products(session, base_url)
            if not products:
                logging.error("No products fetched. Exiting.")
                return
            logging.info(f"Fetched {len(products)} products total.")
            global_pbar.update(1)

            # Separate base products and variants
            base_products = [p for p in products if p.get('variantsCount', -1) > 0]
            variants = [p for p in products if p.get('meta', {}).get('type') == 'variant']

            # Process base products first
            results = []
            async def limited_fetch_product_details_wrapper(prod):
                async with semaphore:
                    return await fetch_product_details(session, prod, base_product_paths)

            for chunk in chunkify(base_products, CHUNK_SIZE):
                tasks = [limited_fetch_product_details_wrapper(p) for p in chunk]
                for coro in tqdm(
                    asyncio.as_completed(tasks),
                    total=len(tasks),
                    desc=color.YELLOW + "Fetching base product details" + color.END
                ):
                    try:
                        details = await coro
                        results.append(details)
                    except Exception as ex:
                        logging.error(f"Error in fetch_product_details: {ex}")
                        
            global_pbar.update(1)

            # Process variants
            for chunk in chunkify(variants, CHUNK_SIZE):
                tasks = [limited_fetch_product_details_wrapper(p) for p in chunk]
                for coro in tqdm(
                    asyncio.as_completed(tasks),
                    total=len(tasks),
                    desc=color.YELLOW + "Fetching variant details" + color.END
                ):
                    try:
                        details = await coro
                        results.append(details)
                    except Exception as ex:
                        logging.error(f"Error in fetch_product_details: {ex}")

            logging.info(f"Fetched details for {len(results)} products.")
            
            global_pbar.update(1)

        # 3) Convert raw results to a DataFrame
        out_data = []
        for r in results:
            if not r:
                continue
            base_data = {
                'Путь': r['path'],
                'Наименование': r['name'],
                'Категория': r['category'],
                'Код товара': r['code'],
                'Порядковый номер': None,
                'Включено в план размещения': "-",
                'Фото на серевере': "-",
                'Дней на складе': r['days'],
                'Остаток': r['stock'],
                'ID': r['id']
            }
            # Add each relevant price
            for price_name, price_value in r['prices'].items():
                base_data[price_name] = price_value
            out_data.append(base_data)

        df_current = pd.DataFrame(out_data)
        global_pbar.update(1)

        # 4) Compare to previous CSV (if exists), to track changes
        combined_data = compare_with_previous_run(df_current, previous_csv)
        global_pbar.update(1)

        # 5) Write or append a sheet in the Excel file
        timestamp = datetime.now().strftime("Export_%Y%m%d_%H%M%S")

        try:
            if os.path.exists(filename):
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                    combined_data.to_excel(writer, sheet_name=timestamp, index=False)
            else:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    combined_data.to_excel(writer, sheet_name=timestamp, index=False)

            print(
                color.GREEN +
                f"Data saved into {filename}, sheet name: {timestamp}" +
                color.END
            )
        except Exception as e:
            logging.error(f"Failed to write to Excel: {e}")
            return
        global_pbar.update(1)

        # 6) Save current subset (Код товара, Остаток) to CSV for next comparison
        df_current[['Код товара', 'Остаток']].to_csv(previous_csv, index=False)
        global_pbar.update(1)

        # 7) Highlight changes in the newly created sheet
        add_dropdown_and_formatting(filename, sheet_name=timestamp)
        global_pbar.update(1)
        wb = openpyxl.load_workbook(filename)
        ws = wb[timestamp]
        highlight_changes(ws, combined_data)
        wb.save(filename)

        # 8) Save data to the database
        update_database(df_current, db_path)
        global_pbar.update(1)
        

        # 9) Optionally, open the Excel file (Windows only)
        try:
            os.startfile(filename)
        except Exception as e:
            logging.error(f"Could not open file {filename}: {e}")

    # ------------------------------------------------------------------------------
    # Script entry point
    # ------------------------------------------------------------------------------
if __name__ == '__main__':
    try:
        asyncio.run(main())
    except Exception as ex:
        logging.error(f"Unhandled exception: {ex}")
