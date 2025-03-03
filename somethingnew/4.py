import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import aiohttp
import asyncio
import pandas as pd
from datetime import datetime
from requests.auth import HTTPBasicAuth
import logging
from aiohttp import BasicAuth, ClientSession, ClientTimeout
from tqdm.asyncio import tqdm
import os
import getpass
from typing import Any, Dict, Optional, List
import psutil
import sqlite3
import smtplib
from email.mime.text import MIMEText
import traceback

# -------------------------------------------------------------------------------
# Logging Setup
# -------------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s',
    handlers=[logging.FileHandler('app.log', mode='w')]
)

# -------------------------------------------------------------------------------
# Simple color-coded print helper for console messages
# -------------------------------------------------------------------------------
class color:
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

# -------------------------------------------------------------------------------
# Compare current and previous runs to flag changes
# -------------------------------------------------------------------------------
def compare_with_previous_run(
    current_data: pd.DataFrame,
    previous_file: str
) -> pd.DataFrame:
    # If no previous CSV exists, mark no changes
    if not os.path.exists(previous_file):
        current_data['Change'] = ''
        return current_data

    # Load previous run data from CSV
    previous_data = pd.read_csv(previous_file, dtype={'Код товара': str})

    # Ensure key columns are strings for proper merging
    current_data['Код товара'] = current_data['Код товара'].astype(str)
    previous_data['Код товара'] = previous_data['Код товара'].astype(str)

    # Initialize 'Change' column as empty
    current_data['Change'] = ''
    previous_data['Change'] = ''

    # Mark new rows that are not present in the previous run
    new_mask = ~current_data['Код товара'].isin(previous_data['Код товара'])
    current_data.loc[new_mask, 'Change'] = 'New'

    # Identify disappeared rows and mark them
    disappeared_mask = ~previous_data['Код товара'].isin(current_data['Код товара'])
    disappeared_rows = previous_data.loc[disappeared_mask].copy()
    disappeared_rows['Change'] = 'Disappeared'

    # Identify rows with changed stock amounts
    merged_data = pd.merge(
        current_data[['Код товара', 'Остаток']],
        previous_data[['Код товара', 'Остаток']],
        on='Код товара',
        how='inner',
        suffixes=('_current', '_previous')
    )
    changed_mask = merged_data['Остаток_current'] != merged_data['Остаток_previous']
    changed_codes = merged_data.loc[changed_mask, 'Код товара'].unique()
    current_data.loc[current_data['Код товара'].isin(changed_codes), 'Change'] = 'Stock Changed'

    # Combine current data and disappeared rows into one DataFrame
    combined_data = pd.concat([current_data, disappeared_rows], ignore_index=True)
    return combined_data

# -------------------------------------------------------------------------------
# Apply conditional formatting to highlight changes in the Excel sheet
# -------------------------------------------------------------------------------
def highlight_changes(ws, combined_data: pd.DataFrame):
    """
    Applies color highlighting in the Excel worksheet based on change status.
    """
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # For new rows
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # For disappeared rows
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # For stock changes

    for index, row in combined_data.iterrows():
        change_type = row.get('Change', '')
        if change_type == 'New':
            for cell in ws[index + 2]:
                cell.fill = green_fill
        elif change_type == 'Disappeared':
            for cell in ws[index + 2]:
                cell.fill = red_fill
        elif change_type == 'Stock Changed':
            for cell in ws[index + 2]:
                cell.fill = yellow_fill

# -------------------------------------------------------------------------------
# Format the Excel sheet with dropdowns, conditional formatting, and column widths
# -------------------------------------------------------------------------------
def add_dropdown_and_formatting(file_path: str, sheet_name: str) -> None:
    """
    Opens or creates an Excel workbook, applies formatting and dropdowns,
    and sets up conditional formatting.
    """
    dropdown_columns = ['F', 'G']
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Dynamically determine the last row and last column
    max_row = ws.max_row
    max_column = ws.max_column
    last_column_letter = get_column_letter(max_column)

    # Add dropdown menu for specified columns
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=True)
    ws.add_data_validation(dv)
    for col in dropdown_columns:
        for row in range(2, max_row + 1):
            cell = f'{col}{row}'
            dv.add(ws[cell])

    # Conditional formatting for "Да" (green) and "Нет" (orange)
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(
            f'{col}2:{col}{max_row}',
            CellIsRule(operator='equal', formula=['"Да"'], stopIfTrue=True, fill=green_fill)
        )
        ws.conditional_formatting.add(
            f'{col}2:{col}{max_row}',
            CellIsRule(operator='equal', formula=['"Нет"'], stopIfTrue=True, fill=orange_fill)
        )

    ws.auto_filter.ref = f"A1:{last_column_letter}{max_row}"
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    for col in ['A', 'E', 'F', 'G', 'K']:
        ws.column_dimensions[col].width = 15

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    wb.save(file_path)
    # print(color.CYAN + f"Файл отформатирован {file_path}" + color.END)

# -------------------------------------------------------------------------------
# Remote Log System via Email Notification
# -------------------------------------------------------------------------------
def send_email_notification(subject: str, message: str):
    """
    Sends an email notification with the given subject and message.
    Configure the SMTP settings below.
    """
    # --- Configuration (update these values) ---
    smtp_server = 'smtp.example.com'
    smtp_port = 587
    smtp_user = 'your_email@example.com'
    smtp_password = 'your_email_password'
    recipient = 'your_email@example.com'
    # ----------------------------------------------

    msg = MIMEText(message)
    msg['Subject'] = subject
    msg['From'] = smtp_user
    msg['To'] = recipient

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Secure the connection
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        logging.info("Email notification sent successfully.")
    except Exception as e:
        logging.error(f"Failed to send email notification: {e}")

# -------------------------------------------------------------------------------
# Function to prompt user for credentials with retry limit
# -------------------------------------------------------------------------------
MAX_LOGIN_ATTEMPTS = 3  # Maximum times user can retry

def get_credentials():
    """
    Prompt user for login credentials securely. Limits retries.
    """
    attempts = 0
    while attempts < MAX_LOGIN_ATTEMPTS:
        username = input("🔑 Enter your email: ")
        password = getpass.getpass("🔒 Enter your password: ")
        if username and password:
            return username, password
        print("⚠️ Credentials cannot be empty. Try again.")
        attempts += 1
    print("❌ Too many failed attempts. Exiting...")
    exit(1)

# -------------------------------------------------------------------------------
# Initial Authentication Setup
# -------------------------------------------------------------------------------
USERNAME, PASSWORD = get_credentials()
auth = BasicAuth(USERNAME, PASSWORD)

base_url = "https://api.moysklad.ru/api/remap/1.2/entity/assortment"
MAX_REQUESTS = 5        # Limit concurrent requests
PAGE_SIZE = 1000        # MoySklad max page size
INCLUDED_PRICE_TYPES = [
    "Цена розница",
    "Цена маркетплейс",
    "Цена мелкий опт",
    "Цена средний опт"
]

# -------------------------------------------------------------------------------
# Async function to fetch data with retries and error handling
# -------------------------------------------------------------------------------
async def fetch(
    session: ClientSession,
    url: str,
    retries: int = 5
) -> Optional[Dict[str, Any]]:
    global auth
    login_attempts = 0
    for attempt in range(retries):
        try:
            async with session.get(url, auth=auth) as response:
                if response.status == 401:
                    logging.error(f"❌ Unauthorized access (401) to {url}. Requesting new credentials.")
                    if login_attempts >= MAX_LOGIN_ATTEMPTS:
                        print("❌ Too many incorrect login attempts. Exiting...")
                        exit(1)
                    new_username, new_password = get_credentials()
                    auth = BasicAuth(new_username, new_password)
                    login_attempts += 1
                    logging.info("🔄 Retrying request with new credentials...")
                    continue
                response.raise_for_status()
                return await response.json()
        except aiohttp.ClientResponseError as e:
            logging.error(f"🚨 Request failed ({e.status}) for {url}: {e}")
            return None
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            logging.warning(f"⚠️ Attempt {attempt + 1} failed: {e}")
            if attempt < retries - 1:
                await asyncio.sleep(1)
            else:
                logging.error(f"❌ All {retries} attempts failed for {url}")
                return None
    return None

# -------------------------------------------------------------------------------
# Fetch product details and allow easy extension for additional API data
# -------------------------------------------------------------------------------
async def fetch_product_details(
    session: ClientSession,
    product: Dict[str, Any],
    base_product_paths: Dict[str, str]
) -> Optional[Dict[str, Any]]:
    """
    Fetch detailed product information.
    
    # NOTE: To add more details from the API in the future,
    # you can extend the returned dictionary below by extracting additional
    # fields from the 'product' dict.
    """
    path_name = product.get('pathName', "")

    # If it's a base product, record its path
    if product.get('variantsCount', -1) > 0:
        base_product_paths[product['id']] = path_name

    # If it's a variant, try to get the parent product's path
    if product.get('meta', {}).get('type') == 'variant':
        parent_id = product.get('product', {}).get('meta', {}).get('href', '').split('/')[-1]
        path_name = base_product_paths.get(parent_id, "")

    # Extract category from characteristics; extend here if needed
    category_value = "base"
    characteristics: List[Dict[str, Any]] = product.get('characteristics', [])
    for char in characteristics:
        category_value = str(char.get('value', 'base'))
        break

    # Filter salePrices to include only selected types
    sale_prices = product.get('salePrices', [])
    prices = {
        p.get('priceType', {}).get('name', 'Unknown'): (p.get('value', 0) / 100)
        for p in sale_prices
        if p.get('priceType', {}).get('name') in INCLUDED_PRICE_TYPES
    }

    # ---------------------------
    # Extract barcode information:
    # Look for 'ean13' barcodes; if none, output keys from the first barcode dict.
    barcodes_list = product.get("barcodes", [])
    ean13_codes = [barcode.get("ean13") for barcode in barcodes_list if "ean13" in barcode]
    if ean13_codes:
        barcode_value = ",".join(ean13_codes)
    elif barcodes_list:
        # If no 'ean13' found but barcodes exist, output the key names from the first dictionary.
        barcode_value = ",".join(list(barcodes_list[0].keys()))
    else:
        barcode_value = ""
    # ---------------------------

    return {
        'name': product.get('name'),
        'code': product.get('code'),
        'path': path_name,
        'stock': product.get('stock', 0),
        'days': product.get('stockDays', 0),
        'category': category_value,
        'prices': prices,
        'id': product.get('id'),
        'ean13': barcode_value  # New field with barcode information
    }

# -------------------------------------------------------------------------------
# Fetch all products using pagination with a progress bar
# -------------------------------------------------------------------------------
async def fetch_all_products(session: ClientSession, base_url: str, limit: int = PAGE_SIZE):
    offset = 0
    all_items = []
    base_product_paths = {}
    with tqdm(desc="Fetching Products (batches)", unit="batch", leave=False) as pbar:
        while True:
            url = f"{base_url}?limit={limit}&offset={offset}"
            logging.info(f"Fetching page offset={offset} ...")
            data = await fetch(session, url)
            if not data:
                logging.warning("No data returned, stopping pagination.")
                break
            rows = data.get('rows', [])
            if not rows:
                logging.info("No more rows; pagination complete.")
                break
            all_items.extend(rows)
            offset += limit
            pbar.update(1)
    return all_items, base_product_paths

# -------------------------------------------------------------------------------
# Check if the Excel file is open and prompt the user to close it
# -------------------------------------------------------------------------------
def check_and_prompt_close_excel(filename: str):
    while True:
        file_closed = True
        for proc in psutil.process_iter(['pid', 'name']):
            if proc.info['name'] == 'EXCEL.EXE':
                for open_file in proc.open_files():
                    if open_file.path == os.path.abspath(filename):
                        file_closed = False
                        print(color.RED + f"The file {filename} is currently open. Please close it." + color.END)
                        input("Press Enter after closing the file...")
                        break
            if not file_closed:
                break
        if file_closed:
            break

# -------------------------------------------------------------------------------
# Save data into a SQLite database
# -------------------------------------------------------------------------------
def update_database(data: pd.DataFrame, db_path: str):
    conn = sqlite3.connect(db_path)
    data.to_sql('products', conn, if_exists='replace', index=False)
    conn.close()
    # print(color.GREEN + f"Data saved into database {db_path}" + color.END)

# -------------------------------------------------------------------------------
# Main asynchronous routine that performs all steps with progress reporting
# -------------------------------------------------------------------------------
async def main():
    filename = "all_products.xlsx"
    db_path = "all_products.db"
    previous_csv = "last.csv"

    overall_steps = 9  # Total number of major steps

    check_and_prompt_close_excel(filename)

    timeout = ClientTimeout(total=120)
    semaphore = asyncio.Semaphore(MAX_REQUESTS)

    async with ClientSession(timeout=timeout) as session:
        with tqdm(total=overall_steps, desc="Overall Progress", unit="step") as global_pbar:
            # Step 1: Fetch all products
            logging.info("Fetching list of all products...")
            products, base_product_paths = await fetch_all_products(session, base_url)
            if not products:
                logging.error("No products fetched. Exiting.")
                return
            logging.info(f"Fetched {len(products)} products total.")
            global_pbar.update(1)

            # Step 2: Process base products without chunks
            base_products = [p for p in products if p.get('variantsCount', -1) > 0]
            results = []
            async def limited_fetch_product_details_wrapper(prod):
                async with semaphore:
                    return await fetch_product_details(session, prod, base_product_paths)
            tasks = [limited_fetch_product_details_wrapper(p) for p in base_products]
            for coro in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc=color.YELLOW + "Fetching base product details" + color.END, leave=False):
                try:
                    details = await coro
                    results.append(details)
                except Exception as ex:
                    logging.error(f"Error in fetch_product_details: {ex}")
            global_pbar.update(1)

            # Step 3: Process variant products without chunks
            variants = [p for p in products if p.get('meta', {}).get('type') == 'variant']
            tasks = [limited_fetch_product_details_wrapper(p) for p in variants]
            for coro in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc=color.YELLOW + "Fetching variant details" + color.END, leave=False):
                try:
                    details = await coro
                    results.append(details)
                except Exception as ex:
                    logging.error(f"Error in fetch_product_details: {ex}")
            global_pbar.update(1)

            # Step 4: Convert raw results to DataFrame
            out_data = []
            for r in results:
                if not r:
                    continue
                base_data = {
                    'Путь': r['path'],
                    'Наименование': r['name'],
                    'Категория': r['category'],
                    'Код товара': r['code'],
                    'Порядковый номер': None,
                    'Включено в план размещения': "-",
                    'Фото на серевере': "-",
                    'Дней на складе': r['days'],
                    'Остаток': r['stock'],
                    'ID': r['id'],
                    'EAN13': r['ean13']  # New column for barcode data
                }
                for price_name, price_value in r['prices'].items():
                    base_data[price_name] = price_value
                out_data.append(base_data)
            df_current = pd.DataFrame(out_data)
            global_pbar.update(1)

            # Step 5: Compare with previous CSV run to detect changes
            combined_data = compare_with_previous_run(df_current, previous_csv)
            global_pbar.update(1)

            # Step 6: Write new data into Excel file using "current" and "previous" sheets
            if os.path.exists(filename):
                wb = openpyxl.load_workbook(filename)
                if "previous" in wb.sheetnames:
                    ws_prev = wb["previous"]
                    wb.remove(ws_prev)
                if "current" in wb.sheetnames:
                    ws_current = wb["current"]
                    ws_current.title = "previous"
                wb.save(filename)
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a' if os.path.exists(filename) else 'w') as writer:
                df_current.to_excel(writer, sheet_name="current", index=False)
            global_pbar.update(1)

            # Step 7: Save current subset for next run comparison
            df_current[['Код товара', 'Остаток']].to_csv(previous_csv, index=False)
            global_pbar.update(1)

            # Step 8: Apply formatting and highlight changes on the new "current" sheet
            add_dropdown_and_formatting(filename, sheet_name="current")
            wb = openpyxl.load_workbook(filename)
            ws = wb["current"]
            highlight_changes(ws, combined_data)
            wb.save(filename)
            global_pbar.update(1)

            # Step 9: Update the SQLite database with current data
            update_database(df_current, db_path)
            global_pbar.update(1)
            
            print(color.GREEN + f"Data saved into {filename}, sheet name: current" + color.END)
            logging.error(f"All steps completed successfully at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


    try:
        os.startfile(filename)
    except Exception as e:
        logging.error(f"Could not open file {filename}: {e}")

    # send_email_notification(
    #     subject="MoySklad API: Run Completed",
    #     message=f"Data processing completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    # )

# -------------------------------------------------------------------------------
# Function to run the main routine once a day (commented out)
# -------------------------------------------------------------------------------
# To run this script automatically once a day, you could uncomment the run_daily() function,
# and then schedule this script (e.g., using cron on Linux or Task Scheduler on Windows).
# Alternatively, use a robust scheduler like APScheduler.
#
# def run_daily():
#     while True:
#         asyncio.run(main())
#         # Sleep for 24 hours (86400 seconds)
#         time.sleep(86400)
#
# Uncomment the following line to run the scheduler:
# run_daily()

# -------------------------------------------------------------------------------
# Script entry point
# -------------------------------------------------------------------------------
if __name__ == '__main__':
    try:
        asyncio.run(main())
    except Exception as ex:
        logging.error(f"Unhandled exception: {ex}")
        logging.error(traceback.format_exc())
