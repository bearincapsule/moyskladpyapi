import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import aiohttp
import asyncio
import pandas as pd
from datetime import datetime
from requests.auth import HTTPBasicAuth
import logging
from aiohttp import BasicAuth, ClientSession, ClientTimeout
from tqdm.asyncio import tqdm
import os
import getpass
from typing import Any, Dict, Optional, List

# Configure logging to log only to a file
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('app.log', mode='w')]
)

class color:
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

print(color.BOLD + color.RED + 'ПОЖАЛУЙСТА, закройте файл product.xlsx, если он запущен, я еще маленькая програмка не умею сама проверять' + color.END)

def compare_with_previous_run(current_data: pd.DataFrame, previous_file: str) -> pd.DataFrame:
    """
    Compares the current data with the data from the previous run stored in a CSV file.
    Highlights rows that are new, disappeared, or have changed stock values.
    """
    if not os.path.exists(previous_file):
        return current_data

    previous_data = pd.read_csv(previous_file)

    # Ensure both columns are of the same type
    current_data['Код товара'] = current_data['Код товара'].astype(str)
    previous_data['Код товара'] = previous_data['Код товара'].astype(str)

    # Find new rows
    new_rows = current_data[~current_data['Код товара'].isin(previous_data['Код товара'])]
    new_rows['Change'] = 'New'

    # Find disappeared rows
    disappeared_rows = previous_data[~previous_data['Код товара'].isin(current_data['Код товара'])]
    disappeared_rows['Change'] = 'Disappeared'

    # Find rows with changed stock values
    merged_data = pd.merge(current_data, previous_data, on='Код товара', suffixes=('_current', '_previous'))
    changed_stock_rows = merged_data[merged_data['Остаток_current'] != merged_data['Остаток_previous']]
    changed_stock_rows['Change'] = 'Stock Changed'

    # Combine the data
    combined_data = pd.concat([current_data, new_rows, disappeared_rows, changed_stock_rows], ignore_index=True)

    return combined_data

def highlight_changes(ws, combined_data):
    """
    Highlights changes in the Excel worksheet.
    """
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for index, row in combined_data.iterrows():
        if row.get('Change') == 'New':
            for cell in ws[index + 2]:
                cell.fill = green_fill
        elif row.get('Change') == 'Disappeared':
            for cell in ws[index + 2]:
                cell.fill = red_fill
        elif row.get('Change') == 'Stock Changed':
            for cell in ws[index + 2]:
                cell.fill = yellow_fill

# Prompt user for authentication details
USERNAME = "warehousetseh2@gmail.com" #input("Enter your username: ")
PASSWORD = "PurestSklad6632!" #getpass.getpass("Enter your password: ")
auth = BasicAuth(USERNAME, PASSWORD)
base_url = "https://api.moysklad.ru/api/remap/1.2/entity/assortment?filter=stockMode=positiveOnly"
included_price_types = ["Цена розница", "Цена маркетплейс", "Цена мелкий опт", "Цена средний опт"]

# Define a global semaphore to limit concurrent HTTP requests
semaphore = asyncio.Semaphore(5)

async def fetch(session: ClientSession, url: str, retries: int = 5) -> Optional[Dict[str, Any]]:
    """
    Fetches JSON data from a URL using the provided session. Retries on failure with specific handling for 429 errors.
    """
    for attempt in range(retries):
        try:
            async with session.get(url, auth=auth) as response:
                response.raise_for_status()
                return await response.json()
        except aiohttp.ClientResponseError as e:
            if e.status == 429:  # Too Many Requests
                logging.warning(f"Attempt {attempt + 1} for {url} failed with 429: {e}")
                if attempt < retries - 1:
                    retry_after = int(e.headers.get('Retry-After', 1))
                    logging.info(f"Retrying after {retry_after} seconds...")
                    await asyncio.sleep(retry_after)
                else:
                    logging.error(f"All {retries} attempts failed for URL: {url}")
                    return None
            else:
                logging.error(f"Request failed for {url}: {e}")
                return None
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            logging.warning(f"Attempt {attempt + 1} for {url} failed: {e}")
            if attempt < retries - 1:
                await asyncio.sleep(1)  # Fixed delay between retries
            else:
                logging.error(f"All {retries} attempts failed for URL: {url}")
                return None
    return None

async def fetch_product_details(session: ClientSession, product: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Fetches detailed information for a product and returns a structured dictionary with relevant data.
    """
    path:str = "fu"
    if product.get('meta', {}).get('type') == 'variant':
        prodvar_req = await fetch(session, product.get('product', {}).get('meta', {}).get('href'))
        path = prodvar_req.get('pathName')
    else:
        path = product.get('pathName')

    # Default category value
    category_value = "base"
    
    # Access characteristics
    characteristics: List[Dict[str, Any]] = product.get('characteristics', [])
    for char in characteristics:
        if char.get('name') == "Категория":
            category_value = char.get('value')
            break  # Stop once category is found

    # Get all sale prices and filter based on included price types
    sale_prices = product.get('salePrices', [])
    prices = {
        price.get('priceType', {}).get('name', 'Unknown'): price.get('value', 0) / 100
        for price in sale_prices
        if price.get('priceType', {}).get('name', 'Unknown') in included_price_types
    }

    return {
        'name': product.get('name'),
        'code': product.get('code'),
        'path': path,
        'stock': product.get('stock', 0),
        'days': product.get('stockDays', 0),
        'category': category_value,
        'prices': prices
    }

async def fetch_all_products(session: ClientSession, base_url: str, limit: int = 1000) -> List[Dict[str, Any]]:
    """
    Fetches all products using pagination.
    """
    offset = 0
    all_products = []
    while True:
        url = f"{base_url}&limit={limit}&offset={offset}"
        response = await fetch(session, url)
        if response is None:
            logging.error("Failed to fetch products or received empty response.")
            break
        products = response.get('rows', [])
        if not products:
            break
        all_products.extend(products)
        offset += limit
    return all_products

async def main() -> None:
    filename = "products.xlsx"
    previous_csv = "last.csv"
    # Set a timeout for the entire session to avoid hanging requests
    timeout = ClientTimeout(total=30)

    async with ClientSession(timeout=timeout) as session:
        logging.info("Fetching all products...")
        products = await fetch_all_products(session, base_url)
        if not products:
            logging.error("Failed to fetch initial product data.")
            return
        logging.info(f"Fetched {len(products)} products.")

        async def limited_fetch_product_details(product: Dict[str, Any]) -> Optional[Dict[str, Any]]:
            async with semaphore:
                return await fetch_product_details(session, product)

        tasks = [limited_fetch_product_details(product) for product in products]

        # Use tqdm to display progress as tasks complete
        results = []
        logging.info("Fetching product details...")
        for result in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc=color.YELLOW + "Загружаю данные" + color.END):
            try:
                details = await result
                results.append(details)
                logging.info(f"Fetched details for product: {details['name'] if details else 'Unknown'}")
            except Exception as e:
                logging.error(f"Error fetching product details: {e}")

        logging.info("Fetched product details.")

        # Prepare data for export
        data = []
        for result in results:
            if result is None:
                continue
            base_data = {
                'Путь': result['path'],
                'Наименование': result['name'],
                'Категория': result['category'],
                'Код товара': result['code'],
                'Порядковый номер': None,
                'Включено в план размещения': "-",
                'Фото на серевере': "-",
                'Дней на складе': result['days'],
                'Остаток': result['stock']
            }
            for price_name, price_value in result['prices'].items():
                base_data[price_name] = price_value
            data.append(base_data)

        # Create a DataFrame and export to Excel
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("Export_%Y%m%d_%H%M%S")

        # Compare with previous run
        combined_data = compare_with_previous_run(df, previous_csv)

        try:
            if os.path.exists(filename):
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                    combined_data.to_excel(writer, sheet_name=timestamp, index=False)
                print(color.GREEN + f"Данные сохранены в {filename} в странице {timestamp}" + color.END)
            else:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    combined_data.to_excel(writer, sheet_name=timestamp, index=False)
                print(color.GREEN + f"Данные сохранены в {filename} в страницу {timestamp}" + color.END)
        except Exception as e:
            logging.error(f"Failed to write data to Excel: {e}")
            return

        # Save current data to CSV for future comparison
        df[['Код товара', 'Остаток']].to_csv(previous_csv, index=False)

        # Add formatting and highlight changes
        wb = openpyxl.load_workbook(filename)
        ws = wb[timestamp]
        highlight_changes(ws, combined_data)
        wb.save(filename)

    # Attempt to open the file; wrap in try/except in case of errors (e.g., non-Windows platforms)
    try:
        os.startfile(filename)
    except Exception as e:
        logging.error(f"Failed to open file {filename}: {e}")

# Run the main function with proper exception handling.
if __name__ == '__main__':
    try:
        asyncio.run(main())
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
