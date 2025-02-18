import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import aiohttp
import asyncio
from aiohttp import BasicAuth
import pandas as pd
from datetime import datetime
from requests.auth import HTTPBasicAuth
import logging
from aiohttp import ClientSession, ClientTimeout
from tqdm.asyncio import tqdm

# Configure logging to log to a file
logging.basicConfig(level=logging.INFO, filename='app.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s')

def add_dropdown_and_formatting(file_path, sheet_name='Sheet1', start_row=2, end_row=2000):
    # Load or create workbook
    dropdown_columns = ['F', 'G']
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    ws = wb.active if sheet_name not in wb.sheetnames else wb[sheet_name]
    ws.title = sheet_name

    # Add dropdown menu
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=True)
    ws.add_data_validation(dv)

    # Apply dropdown to specified range
    for col in dropdown_columns:
        for row in range(start_row, end_row + 1):
            cell = f'{col}{row}'
            dv.add(ws[cell])

    # Apply conditional formatting for "Да" (green)
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(f'{col}{start_row}:{col}{end_row}',
                                      CellIsRule(operator='equal', formula=['"Да"'], stopIfTrue=True, fill=green_fill))

    # Apply conditional formatting for "Нет" (orange)
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(f'{col}{start_row}:{col}{end_row}',
                                      CellIsRule(operator='equal', formula=['"Нет"'], stopIfTrue=True, fill=orange_fill))
    
    # Apply filter to line 195
    ws.auto_filter.ref = f"A1:Z{end_row}"
    
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    for col in ['A','E','F','G','K']: ws.column_dimensions[col].width = 15

    # Apply additional formatting
    for row in ws.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Save the workbook
    wb.save(file_path)
    print(f"Dropdowns and formatting applied to {file_path}")

USERNAME = "warehousetseh2@gmail.com"
PASSWORD = "PurestSklad6632!"
auth = BasicAuth(USERNAME, PASSWORD)
base_url = "https://api.moysklad.ru/api/remap/1.2/report/stock/all"
included_price_types = ["Цена розница", "Цена маркетплейс", "Цена мелкий опт", "Цена средний опт"]

async def fetch(session, url, retries=3):
    for attempt in range(retries):
        try:
            async with session.get(url, auth=auth) as response:
                response.raise_for_status()
                return await response.json()
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            logging.warning(f"Attempt {attempt + 1} failed: {e}")
            if attempt < retries - 1:
                await asyncio.sleep(2 ** attempt)  # Exponential backoff
            else:
                logging.error(f"All {retries} attempts failed for URL: {url}")
                return None

async def fetch_product_details(session, product):
    meta_url = product.get('meta', {}).get('href', {})
    await asyncio.sleep(0.005)  # Add a delay of 5ms
    metaid = await fetch(session, meta_url)
    if metaid is None:
        return None

    # Default category value
    category_value = "dNf"
    
    # Access characteristics
    characteristics = metaid.get('characteristics', [])
    for char in characteristics:
        if char.get('name') == "Категория":
            category_value = char.get('value')
            break  # Stop searching once we find the category

    # Get all sales prices and filter based on included price types
    sale_prices = metaid.get('salePrices', [])
    prices = {
        price.get('priceType', {}).get('name', 'Unknown'): price.get('value', 0) / 100  # Convert to proper format if needed
        for price in sale_prices
        if price.get('priceType', {}).get('name', 'Unknown') in included_price_types
    }

    return {
        'name': product.get('name'),
        'code': product.get('code'),
        'path': product.get('folder', {}).get('name'),
        'stock': product.get('stock', 0),
        'days': product.get('stockDays', 0),
        'category': category_value,
        'prices': prices
    }

async def fetch_all_products(session, base_url, limit=1000):
    offset = 0
    all_products = []
    while True:
        url = f"{base_url}?limit={limit}&offset={offset}"
        response = await fetch(session, url)
        if response is None:
            break
        products = response.get('rows', [])
        if not products:
            break
        all_products.extend(products)
        offset += limit
    return all_products

async def main():
    async with aiohttp.ClientSession() as session:
        products = await fetch_all_products(session, base_url)
        if not products:
            logging.error("Failed to fetch initial data")
            return

        semaphore = asyncio.Semaphore(5)  # Limit concurrent requests to 5

        async def limited_fetch_product_details(product):
            async with semaphore:
                return await fetch_product_details(session, product)

        tasks = [limited_fetch_product_details(product) for product in products]

        # Use tqdm to display a progress bar
        results = []
        for result in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Fetching product details"):
            results.append(await result)

        # Prepare data for export
        data = []
        for result in results:
            if result is None:
                continue
            base_data = {
                'Путь': result['path'],
                'Наименование': result['name'],
                'Категория': result['category'],
                'Код товара': result['code'],
                'Порядковый номер': None,
                'Включено в план размещения': "-",
                'Фото на серевере': "-",
                'Дней на складе': result['days'],
                'Остаток': result['stock']
            }
            for price_name, price_value in result['prices'].items():
                base_data[price_name] = price_value
            data.append(base_data)

        # Create a DataFrame and export to Excel
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"products_{timestamp}.xlsx"
        df.to_excel(filename, index=False)

        # Add dropdown and formatting to the exported file
        add_dropdown_and_formatting(filename)

        print(f"Data exported to {filename}")

# Run the main function
asyncio.run(main())
