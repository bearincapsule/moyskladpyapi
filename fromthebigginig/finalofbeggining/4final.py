import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import aiohttp
import asyncio
import pandas as pd
from datetime import datetime
from requests.auth import HTTPBasicAuth
import logging
from aiohttp import BasicAuth, ClientSession, ClientTimeout
from tqdm.asyncio import tqdm
import os
import zipfile
import getpass
from typing import Any, Dict, Optional, List

# Configure logging to log only to a file
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler('app.log', mode='w')]
)

class color:
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'

print(color.BOLD + color.RED + 'ПОЖАЛУЙСТА, закройте файл product.xlsx, если он запущен, я еще маленькая програмка не умею сама проверять' + color.END)

def add_dropdown_and_formatting(file_path: str, sheet_name: str = 'Sheet1', start_row: int = 2, end_row: int = 2000) -> None:
    """
    Opens or creates an Excel workbook, adds dropdown menus, applies conditional formatting,
    sets column widths and general formatting.
    """
    dropdown_columns = ['F', 'G']
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        ws.title = sheet_name

    # Add dropdown menu
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=True)
    ws.add_data_validation(dv)

    # Apply dropdown to specified range
    for col in dropdown_columns:
        for row in range(start_row, end_row + 1):
            cell = f'{col}{row}'
            dv.add(ws[cell])

    # Apply conditional formatting for "Да" (green)
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(
            f'{col}{start_row}:{col}{end_row}',
            CellIsRule(operator='equal', formula=['"Да"'], stopIfTrue=True, fill=green_fill)
        )

    # Apply conditional formatting for "Нет" (orange)
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(
            f'{col}{start_row}:{col}{end_row}',
            CellIsRule(operator='equal', formula=['"Нет"'], stopIfTrue=True, fill=orange_fill)
        )
    
    # Apply auto-filter
    ws.auto_filter.ref = f"A1:Z{end_row}"

    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    for col in ['A', 'E', 'F', 'G', 'K']:
        ws.column_dimensions[col].width = 15

    # Apply additional formatting: general cell style and header style.
    for row in ws.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    wb.save(file_path)
    print(color.CYAN + f"Файл отформатирован {file_path}" + color.END)

def compare_and_highlight_changes(wb: openpyxl.Workbook, current_sheet: str, previous_sheet: str) -> None:
    """
    Compares two sheets in the workbook and highlights rows that are new or disappeared.
    """
    current_ws = wb[current_sheet]
    previous_ws = wb[previous_sheet]

    # Using first cell of each row as key (ensure it is not None)
    current_data = {row[0].value: row for row in current_ws.iter_rows(min_row=2) if row[0].value is not None}
    previous_data = {row[0].value: row for row in previous_ws.iter_rows(min_row=2) if row[0].value is not None}

    new_row_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    disappeared_row_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for key in current_data:
        if key not in previous_data:
            for cell in current_data[key]:
                cell.fill = new_row_fill

    for key in previous_data:
        if key not in current_data:
            for cell in previous_data[key]:
                cell.fill = disappeared_row_fill

# Prompt user for authentication details
USERNAME = input("Enter your username: ")
PASSWORD = getpass.getpass("Enter your password: ")
auth = BasicAuth(USERNAME, PASSWORD)
base_url = "https://api.moysklad.ru/api/remap/1.2/report/stock/all"
included_price_types = ["Цена розница", "Цена маркетплейс", "Цена мелкий опт", "Цена средний опт"]

# Define a global queue to limit concurrent HTTP requests
queue = asyncio.Queue(maxsize=5)

async def worker(session: ClientSession):
    """
    Worker function to process items from the queue.
    """
    while True:
        url, retries, result = await queue.get()
        try:
            result.append(await fetch(session, url, retries))
        finally:
            queue.task_done()

async def fetch(session: ClientSession, url: str, retries: int = 5) -> Optional[Dict[str, Any]]:
    """
    Fetches JSON data from a URL using the provided session. Retries on failure with specific handling for 429 errors.
    """
    for attempt in range(retries):
        try:
            async with session.get(url, auth=auth) as response:
                response.raise_for_status()
                return await response.json()
        except aiohttp.ClientResponseError as e:
            if e.status == 429:  # Too Many Requests
                logging.warning(f"Attempt {attempt + 1} for {url} failed with 429: {e}")
                if attempt < retries - 1:
                    retry_after = int(e.headers.get('Retry-After', 1))
                    logging.info(f"Retrying after {retry_after} seconds...")
                    await asyncio.sleep(retry_after)
                else:
                    logging.error(f"All {retries} attempts failed for URL: {url}")
                    return None
            else:
                logging.error(f"Request failed for {url}: {e}")
                return None
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            logging.warning(f"Attempt {attempt + 1} for {url} failed: {e}")
            if attempt < retries - 1:
                await asyncio.sleep(1)  # Fixed delay between retries
            else:
                logging.error(f"All {retries} attempts failed for URL: {url}")
                return None
    return None

async def fetch_product_details(session: ClientSession, product: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Fetches detailed information for a product and returns a structured dictionary with relevant data.
    """
    meta_url: Optional[str] = product.get('meta', {}).get('href')
    if not meta_url:
        logging.warning(f"No meta URL found for product: {product.get('name')}")
        return None

    await asyncio.sleep(0.005)  # Small delay between requests
    metaid = await fetch(session, meta_url)
    if metaid is None:
        return None

    # Default category value
    category_value = "dNf"
    
    # Access characteristics
    characteristics: List[Dict[str, Any]] = metaid.get('characteristics', [])
    for char in characteristics:
        if char.get('name') == "Категория":
            category_value = char.get('value')
            break  # Stop once category is found

    # Get all sale prices and filter based on included price types
    sale_prices = metaid.get('salePrices', [])
    prices = {
        price.get('priceType', {}).get('name', 'Unknown'): price.get('value', 0) / 100
        for price in sale_prices
        if price.get('priceType', {}).get('name', 'Unknown') in included_price_types
    }

    return {
        'name': product.get('name'),
        'code': product.get('code'),
        'path': (product.get('folder', {}).get('pathName') or '') + '/' + (product.get('folder', {}).get('name') or ''),
        'stock': product.get('stock', 0),
        'days': product.get('stockDays', 0),
        'category': category_value,
        'prices': prices
    }

async def fetch_all_products(session: ClientSession, base_url: str, limit: int = 1000) -> List[Dict[str, Any]]:
    """
    Fetches all products using pagination.
    """
    offset = 0
    all_products = []
    while True:
        url = f"{base_url}?limit={limit}&offset={offset}"
        response = await fetch(session, url)
        if response is None:
            logging.error("Failed to fetch products or received empty response.")
            break
        products = response.get('rows', [])
        if not products:
            break
        all_products.extend(products)
        offset += limit
    return all_products

async def main() -> None:
    filename = "products.xlsx"
    # Set a timeout for the entire session to avoid hanging requests
    timeout = ClientTimeout(total=30)

    async with ClientSession(timeout=timeout) as session:
        products = await fetch_all_products(session, base_url)
        if not products:
            logging.error("Failed to fetch initial product data.")
            return

        # Create a list to store results
        results = []

        # Create worker tasks to process the queue
        workers = [asyncio.create_task(worker(session)) for _ in range(5)]

        # Enqueue tasks with progress bar
        for product in tqdm(products, desc=color.YELLOW + "Enqueuing tasks" + color.END):
            await queue.put((product.get('meta', {}).get('href'), 5, results))

        # Wait until the queue is fully processed with progress bar
        with tqdm(total=queue.qsize(), desc=color.YELLOW + "Processing tasks" + color.END) as pbar:
            while not queue.empty():
                await asyncio.sleep(0.1)
                pbar.update(1)

        # Cancel worker tasks
        for w in workers:
            w.cancel()

        # Prepare data for export
        data = []
        for result in results:
            if result is None:
                continue
            base_data = {
                'Путь': result['path'],
                'Наименование': result['name'],
                'Категория': result['category'],
                'Код товара': result['code'],
                'Порядковый номер': None,
                'Включено в план размещения': "-",
                'Фото на серевере': "-",
                'Дней на складе': result['days'],
                'Остаток': result['stock']
            }
            for price_name, price_value in result['prices'].items():
                base_data[price_name] = price_value
            data.append(base_data)

        # Create a DataFrame and export to Excel
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("Export_%Y%m%d_%H%M%S")

        try:
            if os.path.exists(filename):
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                    df.to_excel(writer, sheet_name=timestamp, index=False)
                print(color.GREEN + f"Данные сохранены в {filename} в странице {timestamp}" + color.END)
            else:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=timestamp, index=False)
                print(color.GREEN + f"Данные сохранены в {filename} в страницу {timestamp}" + color.END)
        except Exception as e:
            logging.error(f"Failed to write data to Excel: {e}")
            return

        # Add formatting and compare changes between sheets if possible.
        if os.path.exists(filename):
            try:
                add_dropdown_and_formatting(filename, sheet_name=timestamp)
                wb = openpyxl.load_workbook(filename)
                sheet_names = wb.sheetnames
                if len(sheet_names) > 1:
                    compare_and_highlight_changes(wb, current_sheet=sheet_names[-1], previous_sheet=sheet_names[-2])
                wb.save(filename)
            except zipfile.BadZipFile:
                print(f"Error: The file {filename} is not a valid zip file.")
            except Exception as e:
                logging.error(f"Error processing Excel file: {e}")
        else:
            print(f"Error: The file {filename} does not exist.")

    # Attempt to open the file; wrap in try/except in case of errors (e.g., non-Windows platforms)
    try:
        os.startfile(filename)
    except Exception as e:
        logging.error(f"Failed to open file {filename}: {e}")

# Run the main function with proper exception handling.
if __name__ == '__main__':
    try:
        asyncio.run(main())
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
