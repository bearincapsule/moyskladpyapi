import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
import aiohttp
import asyncio
import pandas as pd
from datetime import datetime
from requests.auth import HTTPBasicAuth
import logging
from aiohttp import BasicAuth, ClientSession, ClientTimeout
from tqdm.asyncio import tqdm
import os
import zipfile
import getpass

# Configure logging to log only to a file
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler('app.log', mode='w')
                    ])

class color:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'

print(color.BOLD + color.RED + 'ПОЖАЛУЙСТА, закройте файл product.xlsx, если он запущен, я еще маленькая програмка не умею сама проверять' + color.END)

def add_dropdown_and_formatting(file_path, sheet_name='Sheet1', start_row=2, end_row=2000):
    # Load or create workbook
    dropdown_columns = ['F', 'G']
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    ws = wb.active if sheet_name not in wb.sheetnames else wb[sheet_name]
    ws.title = sheet_name

    # Add dropdown menu
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=True)
    ws.add_data_validation(dv)

    # Apply dropdown to specified range
    for col in dropdown_columns:
        for row in range(start_row, end_row + 1):
            cell = f'{col}{row}'
            dv.add(ws[cell])

    # Apply conditional formatting for "Да" (green)
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(f'{col}{start_row}:{col}{end_row}',
                                      CellIsRule(operator='equal', formula=['"Да"'], stopIfTrue=True, fill=green_fill))

    # Apply conditional formatting for "Нет" (orange)
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for col in dropdown_columns:
        ws.conditional_formatting.add(f'{col}{start_row}:{col}{end_row}',
                                      CellIsRule(operator='equal', formula=['"Нет"'], stopIfTrue=True, fill=orange_fill))
    
    # Apply filter to line 195
    ws.auto_filter.ref = f"A1:Z{end_row}"

    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    for col in ['A','E','F','G','K']: ws.column_dimensions[col].width = 15

    # Apply additional formatting
    for row in ws.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Save the workbook
    wb.save(file_path)
    print(color.CYAN + f"Файл отформатирован {file_path}" + color.END)

def compare_and_highlight_changes(wb, current_sheet, previous_sheet):
    current_ws = wb[current_sheet]
    previous_ws = wb[previous_sheet]

    current_data = {row[0].value: row for row in current_ws.iter_rows(min_row=2)}
    previous_data = {row[0].value: row for row in previous_ws.iter_rows(min_row=2)}

    new_row_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    disappeared_row_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for key in current_data:
        if key not in previous_data:
            for cell in current_data[key]:
                cell.fill = new_row_fill

    for key in previous_data:
        if key not in current_data:
            for cell in previous_data[key]:
                cell.fill = disappeared_row_fill

# Prompt user for authentication details
USERNAME = input("Enter your username: ")
PASSWORD = getpass.getpass("Enter your password: ")
auth = BasicAuth(USERNAME, PASSWORD)
base_url = "https://api.moysklad.ru/api/remap/1.2/report/stock/all"
included_price_types = ["Цена розница", "Цена маркетплейс", "Цена мелкий опт", "Цена средний опт"]

# Define a global semaphore
semaphore = asyncio.Semaphore(5)

async def fetch(session, url, retries=5):
    for attempt in range(retries):
        try:
            async with session.get(url, auth=auth) as response:
                response.raise_for_status()
                return await response.json()
        except aiohttp.ClientResponseError as e:
            if e.status == 429:  # Too Many Requests
                logging.warning(f"Attempt {attempt + 1} failed: {e}")
                if attempt < retries - 1:
                    retry_after = int(response.headers.get('Retry-After', 1))
                    logging.info(f"Retrying after {retry_after} seconds...")
                    await asyncio.sleep(retry_after)
                else:
                    logging.error(f"All {retries} attempts failed for URL: {url}")
                    return None
            else:
                logging.error(f"Request failed: {e}")
                return None
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            logging.warning(f"Attempt {attempt + 1} failed: {e}")
            if attempt < retries - 1:
                await asyncio.sleep(1)  # Fixed delay
            else:
                logging.error(f"All {retries} attempts failed for URL: {url}")
                return None

async def fetch_product_details(session, product):
    meta_url = product.get('meta', {}).get('href', {})
    await asyncio.sleep(0.005)  # Add a delay of 5ms
    metaid = await fetch(session, meta_url)
    if metaid is None:
        return None

    # Default category value
    category_value = "dNf"
    
    # Access characteristics
    characteristics = metaid.get('characteristics', [])
    for char in characteristics:
        if char.get('name') == "Категория":
            category_value = char.get('value')
            break  # Stop searching once we find the category

    # Get all sales prices and filter based on included price types
    sale_prices = metaid.get('salePrices', [])
    prices = {
        price.get('priceType', {}).get('name', 'Unknown'): price.get('value', 0) / 100  # Convert to proper format if needed
        for price in sale_prices
        if price.get('priceType', {}).get('name', 'Unknown') in included_price_types
    }

    return {
        'name': product.get('name'),
        'code': product.get('code'),
        'path': (product.get('folder', {}).get('pathName') or '') + '/' + (product.get('folder', {}).get('name') or ''),
        'stock': product.get('stock', 0),
        'days': product.get('stockDays', 0),
        'category': category_value,
        'prices': prices
    }

async def fetch_all_products(session, base_url, limit=1000):
    offset = 0
    all_products = []
    while True:
        url = f"{base_url}?limit={limit}&offset={offset}"
        response = await fetch(session, url)
        if response is None:
            break
        products = response.get('rows', [])
        if not products:
            break
        all_products.extend(products)
        offset += limit
    return all_products

async def main():
    filename = "products.xlsx"

    async with aiohttp.ClientSession() as session:
        products = await fetch_all_products(session, base_url)
        if not products:
            logging.error("Failed to fetch initial data")
            return

        async def limited_fetch_product_details(product):
            async with semaphore:
                return await fetch_product_details(session, product)

        tasks = [limited_fetch_product_details(product) for product in products]

        # Use tqdm to display a progress bar
        results = []
        for result in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc=color.YELLOW + "Загружаю данные" + color.END):
            results.append(await result)

        # Prepare data for export
        data = []
        for result in results:
            if result is None:
                continue
            base_data = {
                'Путь': result['path'],
                'Наименование': result['name'],
                'Категория': result['category'],
                'Код товара': result['code'],
                'Порядковый номер': None,
                'Включено в план размещения': "-",
                'Фото на серевере': "-",
                'Дней на складе': result['days'],
                'Остаток': result['stock']
            }
            for price_name, price_value in result['prices'].items():
                base_data[price_name] = price_value
            data.append(base_data)

        # Create a DataFrame and export to Excel
        df = pd.DataFrame(data)
        timestamp = datetime.now().strftime("Export_%Y%m%d_%H%M%S")

        # Check if the file exists and load it, otherwise create a new one
        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=timestamp, index=False)
                print(color.GREEN + f"Данные сохранены в {filename} в странице {timestamp}" + color.END)
        else:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=timestamp, index=False)
                print(color.GREEN + f"Данные сохранены в {filename} в страницу {timestamp}" + color.END)

        # Check if the file exists and is a valid Excel file
        if os.path.exists(filename):
            try:
                add_dropdown_and_formatting(filename, sheet_name=timestamp)
                wb = openpyxl.load_workbook(filename)
                sheet_names = wb.sheetnames
                if len(sheet_names) > 1:
                    compare_and_highlight_changes(wb, current_sheet=sheet_names[-1], previous_sheet=sheet_names[-2])
                wb.save(filename)
            except zipfile.BadZipFile:
                print(f"Error: The file {filename} is not a valid zip file.")
        else:
            print(f"Error: The file {filename} does not exist.")

    # Open the file at the completion
    os.startfile(filename)

# Run the main function
asyncio.run(main())
